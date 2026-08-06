"""엑셀 **모델**에 공시 실적을 채워 넣는다.

왜 필요한가
-----------
RA 업무의 중심에 모델 관리가 있고([D55](../../../docs/decisions.md#d55)),
새 공시가 나오면 그 숫자를 모델에 옮겨 적는 것이 반복 노동이다. 우리는 그
숫자를 **검산까지 마친 채로** 갖고 있다.
[Daloopa](https://daloopa.com/benefits/model-updates)가 하는 일이 이것이다 —
**모델을 대체하지 않고 먹인다.**

남의 파일에 쓰는 일이다 — 안전 규칙 넷
--------------------------------------
1. **수식 셀은 절대 건드리지 않는다.** 값이 `=`로 시작하면 손대지 않는다.
   RA의 모델은 수식 덩어리이고, 하나 깨뜨리면 그 사람 하루가 날아간다.
2. **원본을 고치지 않는다.** 사본을 만들어 돌려준다.
3. **무엇을 어디에 썼는지 전부 돌려준다.** 셀 주소와 이전/이후 값을 남겨
   사람이 훑을 수 있어야 한다.
4. **모르면 안 쓴다.** 라벨이 확실히 안 맞으면 건너뛰고 목록에 남긴다.

실측으로 확인한 것
------------------
`openpyxl`로 열고 다시 저장해도 **수식·번호서식·차트·조건부서식·틀고정이
모두 보존**된다. 값 셀만 바꾸면 나머지는 그대로다.

찾는 방법
---------
모델의 **행 라벨**(보통 A열 또는 B열)과 **열 머리행의 연도**를 읽어 격자를
만든다. 우리 지표 이름과 라벨을 정규화해 맞춘다. 하우스마다 표기가 달라
(`매출액` · `매출` · `Revenue` · `매출액(십억원)`) 별칭을 둔다.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

# 우리 지표 → 모델에서 쓰일 법한 라벨들. **좁게 유지한다** — 애매하면 안 쓴다.
ALIASES: dict[str, tuple[str, ...]] = {
    "revenue": ("매출액", "매출", "영업수익", "수익", "revenue", "sales", "netsales"),
    "cost_of_sales": ("매출원가", "costofsales", "cogs"),
    "gross_profit": ("매출총이익", "grossprofit"),
    "sga": ("판매비와관리비", "판관비", "sga", "sgna"),
    "operating_income": ("영업이익", "영업손익", "operatingprofit", "operatingincome", "op"),
    "pretax_income": ("법인세차감전순이익", "세전이익", "세전계속사업이익", "pretaxincome", "ebt"),
    "net_income": ("당기순이익", "순이익", "netincome", "netprofit"),
    "net_income_parent": ("지배주주순이익", "지배주주지분순이익", "지배기업소유주지분순이익"),
    "total_assets": ("자산총계", "총자산", "totalassets"),
    "total_liabilities": ("부채총계", "총부채", "totalliabilities"),
    "total_equity": ("자본총계", "총자본", "totalequity"),
    "cfo": ("영업활동현금흐름", "영업현금흐름", "operatingcashflow", "cfo"),
    "capex": ("capex", "설비투자", "유형자산취득"),
    "ebitda": ("ebitda",),
}

# 라벨 정규화 — 단위 표기·공백·괄호를 떼면 하우스가 달라도 맞는다
_STRIP = re.compile(r"\(.*?\)|[\s,·:\-_/]|십억원|백만원|억원|원|%")
# 열 머리행의 연도. `2025` · `2025A` · `25F` · `FY25`
_YEAR = re.compile(r"(?:FY)?(\d{4}|\d{2})\s*[AFEP]?$", re.IGNORECASE)

MAX_SCAN_ROWS = 400
MAX_SCAN_COLS = 40


def norm(text: object) -> str:
    return _STRIP.sub("", str(text or "")).lower()


_LOOKUP = {alias: metric for metric, names in ALIASES.items() for alias in map(norm, names)}


def _year_of(text: object) -> int | None:
    m = _YEAR.match(str(text or "").strip())
    if m is None:
        return None
    n = int(m.group(1))
    return n if n > 1900 else 2000 + n


@dataclass
class Written:
    """채운 칸 하나."""

    sheet: str
    cell: str
    metric: str
    label: str
    year: int
    before: object
    after: float


@dataclass
class Skipped:
    """안 채운 칸과 그 이유. **조용히 넘기지 않는다.**"""

    sheet: str
    cell: str
    label: str
    year: int
    reason: str


@dataclass
class FillResult:
    data: bytes = b""
    written: list[Written] = field(default_factory=list)
    skipped: list[Skipped] = field(default_factory=list)
    sheets_scanned: list[str] = field(default_factory=list)

    @property
    def usable(self) -> bool:
        return bool(self.written)


def fill_model(
    data: bytes,
    values: dict[str, dict[int, float]],
    *,
    unit: float = 1.0,
) -> FillResult:
    """모델 사본에 값을 채워 돌려준다.

    `values`는 `{지표: {연도: 원 단위 값}}`. `unit`은 모델의 단위 — 백만원
    단위 모델이면 `1_000_000`을 준다(값을 그것으로 나눠 넣는다).

    **수식 셀은 건드리지 않는다.** 값이 `=`로 시작하면 건너뛰고 이유를 남긴다.
    """
    import openpyxl

    result = FillResult()
    wb = openpyxl.load_workbook(io.BytesIO(data))

    for ws in wb.worksheets:
        result.sheets_scanned.append(ws.title)
        rows = min(ws.max_row or 0, MAX_SCAN_ROWS)
        cols = min(ws.max_column or 0, MAX_SCAN_COLS)
        if not rows or not cols:
            continue

        # 연도 열 찾기 — 위쪽 몇 줄에서만 본다. 아래로 갈수록 데이터다.
        year_cols: dict[int, int] = {}
        for r in range(1, min(rows, 12) + 1):
            found = {}
            for c in range(1, cols + 1):
                y = _year_of(ws.cell(row=r, column=c).value)
                if y is not None:
                    found[c] = y
            if len(found) >= 2:  # 한 칸짜리는 연도 머리행이 아니다
                year_cols = found
                break
        if not year_cols:
            continue

        # 행 라벨 찾기 — 왼쪽 세 열까지만 본다
        for r in range(1, rows + 1):
            metric = None
            for c in (1, 2, 3):
                metric = _LOOKUP.get(norm(ws.cell(row=r, column=c).value))
                if metric:
                    break
            if metric is None or metric not in values:
                continue
            label = str(ws.cell(row=r, column=1).value or "").strip()
            for col, year in year_cols.items():
                value = values[metric].get(year)
                if value is None:
                    continue
                cell = ws.cell(row=r, column=col)
                ref = f"{cell.column_letter}{cell.row}"
                # **수식은 절대 건드리지 않는다.**
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    result.skipped.append(Skipped(ws.title, ref, label, year, "수식이 있어 건너뜀"))
                    continue
                if isinstance(cell.value, str) and cell.value.strip():
                    result.skipped.append(
                        Skipped(ws.title, ref, label, year, "숫자가 아닌 값이 있어 건너뜀")
                    )
                    continue
                after = value / unit
                result.written.append(
                    Written(ws.title, ref, metric, label, year, cell.value, after)
                )
                cell.value = after

    buf = io.BytesIO()
    wb.save(buf)
    result.data = buf.getvalue()
    return result
