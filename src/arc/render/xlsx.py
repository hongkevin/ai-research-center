"""노트 → **엑셀(.xlsx)**.

왜 필요한가
-----------
RA 업무의 중심에 **모델 관리**가 있다([D55](../../../docs/decisions.md#d55)).
모델은 엑셀에 살고, 새 공시가 나오면 그 숫자를 모델에 옮겨 적는 것이 반복
노동이다. 우리는 그 숫자를 **검산까지 마친 채로** 갖고 있다.

**숫자를 숫자로 낸다.**
------------------------
노트의 「11조 3,145억원」은 사람이 읽는 문자열이라 모델에서 쓸 수 없다.
그래서 이 모듈은 마크다운이 아니라 **레지스트리를 읽는다** — `value`가
원 단위 정수다. 붙여넣으면 바로 수식에 들어간다.

단위는 **백만원**으로 낸다. 원 단위 정수는 자릿수가 많아 눈으로 읽히지 않고,
국내 리포트 관행이 십억원·백만원이다. 원본 값은 「수치 출처」 시트에 그대로
남겨 되짚을 수 있게 한다.

시트 구성
---------
`요약` · `손익계산서` · `재무상태표` · `현금흐름표` · `부문` · `추정` ·
**`수치 출처`**. 마지막 시트가 이 파일의 차별점이다 — 셀 하나하나가 어느
공시 어느 절에서 왔는지 링크까지 붙는다([D36](../../../docs/decisions.md#d36) ·
[D44](../../../docs/decisions.md#d44)).
"""

from __future__ import annotations

import io
import re
from collections import defaultdict

from arc.finmodel.metrics import (
    BALANCE_SHEET_METRICS,
    CASH_FLOW_METRICS,
    INCOME_STATEMENT_METRICS,
)

MILLION = 1_000_000

# `revenue_2025a` · `operating_income_2026e` → (지표, 연도, 실적/추정)
_KEY = re.compile(r"^(?P<metric>.+?)_(?P<year>\d{4})(?P<kind>[ae])$")

# 비율·배수는 백만원으로 나누면 안 된다.
_RATIO_UNITS = ("%", "pp", "%p", "배")


def _split(key: str) -> tuple[str, int, str] | None:
    m = _KEY.match(key)
    if m is None:
        return None
    return m.group("metric"), int(m.group("year")), m.group("kind")


def _sheet(wb, title: str):
    ws = wb.create_sheet(title)
    ws.freeze_panes = "B2"
    return ws


def _write_matrix(ws, rows: list[tuple[str, dict[int, float], str]], years: list[int]) -> None:
    """행 라벨 × 연도 격자. **셀에 숫자를 넣는다.**"""
    from openpyxl.styles import Font

    ws.cell(row=1, column=1, value="항목").font = Font(bold=True)
    for c, year in enumerate(years, start=2):
        # **연도에 천단위 구분이 붙으면 안 된다** — `2,025`가 된다.
        cell = ws.cell(row=1, column=c, value=year)
        cell.font = Font(bold=True)
        cell.number_format = "0"
    ws.cell(row=1, column=len(years) + 2, value="단위").font = Font(bold=True)

    for r, (label, by_year, unit) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=label)
        ratio = unit in _RATIO_UNITS
        for c, year in enumerate(years, start=2):
            value = by_year.get(year)
            if value is None:
                continue
            cell = ws.cell(row=r, column=c, value=value if ratio else value / MILLION)
            cell.number_format = "0.0" if ratio else "#,##0"
        ws.cell(row=r, column=len(years) + 2, value=unit if ratio else "백만원")

    ws.column_dimensions["A"].width = 22
    for c in range(2, len(years) + 3):
        ws.column_dimensions[chr(64 + c)].width = 15


def _collect(registry: list[dict]) -> tuple[dict, dict, list[int]]:
    """레지스트리 → {지표: {연도: 값}} · {지표: (라벨, 단위)} · 연도 목록."""
    values: dict[str, dict[int, float]] = defaultdict(dict)
    meta: dict[str, tuple[str, str]] = {}
    years: set[int] = set()
    for e in registry:
        if e.get("internal"):
            continue
        got = _split(str(e.get("key", "")))
        if got is None:
            continue
        metric, year, _kind = got
        value = e.get("value")
        if value is None:
            continue
        # 추정 연도는 실적과 겹치지 않으므로 한 격자에 같이 놓을 수 있다
        values[metric][year] = float(value)
        years.add(year)
        label = str(e.get("label") or metric)
        # 라벨 끝의 `(2025A)`를 뗀다 — 열 머리행이 이미 연도를 말한다
        meta.setdefault(
            metric, (re.sub(r"\s*\(\d{4}[AEae]\)\s*$", "", label), str(e.get("unit") or ""))
        )
    return values, meta, sorted(years)


def _rows_for(
    metrics: tuple[str, ...], values: dict, meta: dict
) -> list[tuple[str, dict[int, float], str]]:
    out = []
    for m in metrics:
        if m in values:
            label, unit = meta[m]
            out.append((label, values[m], unit))
    return out


def note_to_xlsx(
    registry: list[dict],
    *,
    company: str = "",
    symbol: str = "",
    market: str = "",
    basis: str = "",
    period_label: str = "",
) -> bytes:
    """레지스트리 → .xlsx 바이트.

    **마크다운이 아니라 레지스트리를 읽는다.** 노트의 「11조 3,145억원」은
    사람이 읽는 문자열이라 모델에서 못 쓴다. 여기서는 원 단위 정수를 백만원으로
    나눠 **숫자 셀**로 넣는다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    values, meta, years = _collect(registry)
    wb = Workbook()
    wb.remove(wb.active)

    # ── 요약 ────────────────────────────────────────────────────────
    ws = wb.create_sheet("요약")
    for r, (k, v) in enumerate(
        (
            ("회사", company),
            ("종목코드", symbol),
            ("시장", market),
            ("기준 보고서", period_label),
            ("회계기준", basis),
            ("단위", "백만원 (비율은 % 그대로)"),
            ("출처", "금융감독원 전자공시시스템(DART)"),
            (
                "주의",
                "이 파일의 수치는 공시에서 읽어 검산한 값입니다. 추정은 가정에 따른 계산입니다.",
            ),
        ),
        start=1,
    ):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 62

    for title, metrics in (
        ("손익계산서", INCOME_STATEMENT_METRICS),
        ("재무상태표", BALANCE_SHEET_METRICS),
        ("현금흐름표", CASH_FLOW_METRICS),
    ):
        rows = _rows_for(metrics, values, meta)
        if rows:
            _write_matrix(_sheet(wb, title), rows, years)

    # ── 부문 ────────────────────────────────────────────────────────
    seg = [m for m in sorted(values) if m.startswith("segment")]
    if seg:
        rows = [(meta[m][0], values[m], meta[m][1]) for m in seg]
        _write_matrix(_sheet(wb, "부문"), rows, years)

    # ── 나머지 지표 (마진·추정 가정 등) ──────────────────────────────
    used = set(INCOME_STATEMENT_METRICS) | set(BALANCE_SHEET_METRICS) | set(CASH_FLOW_METRICS)
    rest = [m for m in sorted(values) if m not in used and not m.startswith("segment")]
    if rest:
        rows = [(meta[m][0], values[m], meta[m][1]) for m in rest]
        _write_matrix(_sheet(wb, "지표·추정"), rows, years)

    # ── 수치 출처 ───────────────────────────────────────────────────
    # **이 시트가 이 파일의 차별점이다.** 셀 하나하나가 어느 공시 어느 절에서
    # 왔는지 링크까지 붙는다. 모델에 붙여 넣고도 되짚을 수 있다.
    ws = wb.create_sheet("수치 출처")
    head = ("키", "항목", "값(원)", "단위", "산식", "출처", "공시", "확인 링크")
    for c, name in enumerate(head, start=1):
        ws.cell(row=1, column=c, value=name).font = Font(bold=True)
    r = 2
    for e in registry:
        if e.get("internal"):
            continue
        prov = e.get("provenance") or {}
        ws.cell(row=r, column=1, value=e.get("key"))
        ws.cell(row=r, column=2, value=e.get("label"))
        ws.cell(row=r, column=3, value=e.get("value")).number_format = "#,##0"
        ws.cell(row=r, column=4, value=e.get("unit"))
        ws.cell(row=r, column=5, value=e.get("formula"))
        ws.cell(row=r, column=6, value=prov.get("dataset"))
        ws.cell(row=r, column=7, value=prov.get("source_ref"))
        url = prov.get("verify_url")
        if url:
            cell = ws.cell(row=r, column=8, value="열기")
            cell.hyperlink = url
            cell.style = "Hyperlink"
        r += 1
    for col, width in zip("ABCDEFGH", (26, 30, 18, 8, 34, 34, 18, 10), strict=False):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
