"""엑셀 모델에 공시 실적을 채워 넣기.

**남의 파일에 쓰는 일이다.** RA의 모델은 수식 덩어리이고, 하나 깨뜨리면 그
사람 하루가 날아간다. 이 파일이 지키는 것은 그 안전 규칙 넷이다:

1. 수식 셀은 절대 건드리지 않는다
2. 원본을 고치지 않고 사본을 준다
3. 무엇을 어디에 썼는지 전부 돌려준다
4. 모르면 안 쓴다
"""

from __future__ import annotations

import io

import openpyxl

from arc.ingest.model_fill import fill_model, norm

BILLION = 1_000_000_000


def _model(*, years=("2023A", "2024A", "2025A"), labels=("매출액", "영업이익")) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "손익"
    ws["A1"] = "(십억원)"
    for i, y in enumerate(years, start=2):
        ws.cell(row=1, column=i, value=y)
    for r, label in enumerate(labels, start=2):
        ws.cell(row=r, column=1, value=label)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet(data: bytes, name="손익"):
    return openpyxl.load_workbook(io.BytesIO(data))[name]


VALUES = {
    "revenue": {2025: 11_314 * BILLION},
    "operating_income": {2025: 913 * BILLION},
}


class TestSafety:
    def test_formula_cells_are_never_touched(self):
        """**제일 중요한 규칙.** 수식 하나 깨지면 그 사람 하루가 날아간다."""
        wb = openpyxl.load_workbook(io.BytesIO(_model()))
        ws = wb["손익"]
        ws["D2"] = "=C2*1.1"  # 매출액 2025 자리에 수식
        buf = io.BytesIO()
        wb.save(buf)

        got = fill_model(buf.getvalue(), VALUES, unit=BILLION)
        assert _sheet(got.data)["D2"].value == "=C2*1.1"
        assert any(s.reason == "수식이 있어 건너뜀" for s in got.skipped)

    def test_existing_text_is_not_overwritten(self):
        wb = openpyxl.load_workbook(io.BytesIO(_model()))
        wb["손익"]["D2"] = "N/A"
        buf = io.BytesIO()
        wb.save(buf)

        got = fill_model(buf.getvalue(), VALUES, unit=BILLION)
        assert _sheet(got.data)["D2"].value == "N/A"
        assert got.skipped

    def test_original_bytes_are_not_modified(self):
        """원본을 고치지 않고 사본을 준다."""
        original = _model()
        before = bytes(original)
        fill_model(original, VALUES, unit=BILLION)
        assert original == before

    def test_charts_and_formats_survive(self):
        """차트를 날리면 그게 피해다."""
        from openpyxl.chart import BarChart, Reference

        wb = openpyxl.load_workbook(io.BytesIO(_model()))
        ws = wb["손익"]
        ws["B2"], ws["C2"] = 100, 200
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=2, max_col=3))
        ws.add_chart(chart, "F2")
        ws["B2"].number_format = "#,##0.0"
        ws.freeze_panes = "B2"
        buf = io.BytesIO()
        wb.save(buf)

        out = _sheet(fill_model(buf.getvalue(), VALUES, unit=BILLION).data)
        assert len(out._charts) == 1
        assert out["B2"].number_format == "#,##0.0"
        assert out.freeze_panes == "B2"


class TestFilling:
    def test_writes_into_the_matching_year_column(self):
        got = fill_model(_model(), VALUES, unit=BILLION)
        ws = _sheet(got.data)
        assert ws["D2"].value == 11_314  # 매출액 · 2025A
        assert ws["D3"].value == 913  # 영업이익 · 2025A

    def test_untouched_years_stay_empty(self):
        """2025만 줬으면 2023·2024는 그대로다."""
        ws = _sheet(fill_model(_model(), VALUES, unit=BILLION).data)
        assert ws["B2"].value is None and ws["C2"].value is None

    def test_unit_divides_the_value(self):
        ws = _sheet(fill_model(_model(), VALUES, unit=1_000_000).data)
        assert ws["D2"].value == 11_314_000

    def test_every_write_is_reported(self):
        """파일만 주면 무엇이 바뀌었는지 알 수 없다."""
        got = fill_model(_model(), VALUES, unit=BILLION)
        assert {(w.cell, w.label) for w in got.written} == {("D2", "매출액"), ("D3", "영업이익")}
        assert all(w.before is None for w in got.written)


class TestMatching:
    def test_house_variants_match(self):
        """하우스마다 표기가 다르다."""
        for label in ("매출액", "매출", "Revenue", "매출액(십억원)", "Net Sales"):
            got = fill_model(_model(labels=(label,)), VALUES, unit=BILLION)
            assert got.written, label

    def test_unknown_label_is_left_alone(self):
        """**모르면 안 쓴다.**"""
        got = fill_model(_model(labels=("이상한항목",)), VALUES, unit=BILLION)
        assert not got.written

    def test_two_digit_and_fy_years(self):
        for y in ("25A", "FY25", "2025F"):
            got = fill_model(_model(years=("23A", "24A", y)), VALUES, unit=BILLION)
            assert got.written, y

    def test_sheet_without_a_year_header_is_skipped(self):
        wb = openpyxl.Workbook()
        wb.active["A2"] = "매출액"
        buf = io.BytesIO()
        wb.save(buf)
        assert not fill_model(buf.getvalue(), VALUES, unit=BILLION).written

    def test_normalisation_strips_units_and_spaces(self):
        assert norm("매출액 (십억원)") == norm("매출액") == "매출액"
