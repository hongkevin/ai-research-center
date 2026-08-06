"""노트 → 엑셀.

RA 업무의 중심에 모델 관리가 있고(D55), 모델은 엑셀에 산다. 새 공시가 나오면
그 숫자를 모델에 옮겨 적는 것이 반복 노동이다.

**숫자가 숫자여야 한다.** 노트의 「11조 3,145억원」은 사람이 읽는 문자열이라
모델에서 못 쓴다. 그래서 이 경로는 마크다운이 아니라 레지스트리를 읽는다.
"""

from __future__ import annotations

import io

import openpyxl

from arc.render.xlsx import note_to_xlsx

REG = [
    {
        "key": "revenue_2024a",
        "value": 10_294_103_000_000,
        "unit": "원",
        "label": "매출액 (2024A)",
        "provenance": {"dataset": "재무제표", "source_ref": "2026", "verify_url": "https://x/1"},
    },
    {
        "key": "revenue_2025a",
        "value": 11_314_459_000_000,
        "unit": "원",
        "label": "매출액 (2025A)",
        "provenance": {"dataset": "재무제표", "source_ref": "2026", "verify_url": "https://x/1"},
    },
    {
        "key": "operating_margin_2025a",
        "value": 8.1,
        "unit": "%",
        "label": "영업이익률 (2025A)",
        "provenance": {},
    },
    {
        "key": "total_assets_2025a",
        "value": 14_595_900_000_000,
        "unit": "원",
        "label": "자산총계 (2025A)",
        "provenance": {},
    },
    {
        "key": "cfo_2025a",
        "value": 1_490_091_000_000,
        "unit": "원",
        "label": "영업활동현금흐름 (2025A)",
        "provenance": {},
    },
    # 내부 검산값은 독자에게 소음이다
    {
        "key": "segment_gap_2025a",
        "value": 0.0,
        "unit": "%",
        "label": "부문 검산 차이",
        "internal": True,
        "provenance": {},
    },
]


def _wb(registry=REG):
    return openpyxl.load_workbook(io.BytesIO(note_to_xlsx(registry, company="삼성전기")))


class TestSheets:
    def test_statements_get_their_own_sheets(self):
        assert {"요약", "손익계산서", "재무상태표", "현금흐름표", "수치 출처"} <= set(
            _wb().sheetnames
        )

    def test_empty_registry_still_makes_a_file(self):
        wb = _wb([])
        assert "요약" in wb.sheetnames


class TestNumbers:
    def test_values_are_numbers_not_strings(self):
        """**이게 이 파일의 존재 이유다.** 문자열이면 모델에서 못 쓴다."""
        ws = _wb()["손익계산서"]
        cell = ws.cell(row=2, column=3)  # 매출액 · 2025
        assert isinstance(cell.value, int | float)

    def test_amounts_are_in_millions(self):
        """원 단위 정수는 자릿수가 많아 눈으로 안 읽힌다."""
        ws = _wb()["손익계산서"]
        assert ws.cell(row=2, column=3).value == 11_314_459

    def test_ratios_are_not_divided(self):
        """8.1%를 백만으로 나누면 0이 된다."""
        ws = _wb()["지표·추정"]
        found = [
            ws.cell(row=r, column=c).value
            for r in range(2, ws.max_row + 1)
            for c in range(2, ws.max_column)
            if ws.cell(row=r, column=1).value == "영업이익률"
        ]
        assert 8.1 in found

    def test_year_headers_have_no_thousand_separator(self):
        """`2,025`가 되면 안 된다."""
        ws = _wb()["손익계산서"]
        assert ws.cell(row=1, column=2).number_format == "0"

    def test_label_drops_the_year_suffix(self):
        """열 머리행이 이미 연도를 말한다."""
        assert _wb()["손익계산서"].cell(row=2, column=1).value == "매출액"


class TestProvenance:
    def test_every_number_is_traceable(self):
        """**이 시트가 차별점이다.** 모델에 붙여 넣고도 되짚을 수 있어야 한다."""
        ws = _wb()["수치 출처"]
        keys = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "revenue_2025a" in keys

    def test_verify_link_is_a_hyperlink(self):
        ws = _wb()["수치 출처"]
        row = next(
            r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == "revenue_2025a"
        )
        assert ws.cell(row=row, column=8).hyperlink is not None

    def test_internal_check_values_are_left_out(self):
        """검산값은 감사에는 필요하지만 독자에게는 소음이다."""
        ws = _wb()["수치 출처"]
        keys = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "segment_gap_2025a" not in keys
